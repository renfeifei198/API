from testFile.getPath import get_Path
import openpyxl

'''
用例写入测试结果
'''

path = get_Path()

class writeExcel():
    '''写入测试结果'''

    def write_xls(self,xls_name,sheet_name):
        self.xlsPath = path + '\case_excel\\' + xls_name
        self.file = openpyxl.load_workbook(self.xlsPath)  # 获取excel并打开
        self.sheet_name = self.file.active = self.file[sheet_name] # 获取sheet名并激活

    def write_fail(self,row,col,value):
        '''写入FAIL'''
        self.sheet_name.cell(row=row,column=col,value=value) # 写入值
        self.file.save(self.xlsPath)


    def write_pass(self,row,col,value):
        '''写入PASS'''
        self.sheet_name.cell(row=row,column=col,value=value) # 写入值
        self.file.save(self.xlsPath)


if __name__ == '__main__': # 测试写入

    a = writeExcel()
    a.write_xls('login_case.xlsx','login1')
    a.write_fail(2,6,'PASS')
    a.write_pass(3,6,'FAIL')
